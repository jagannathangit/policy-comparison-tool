#!/usr/bin/env python3
"""
Generate Policy Comparison XLSX Test Data
This script generates a test XLSX file with 100 policy entries for website comparison
"""

import openpyxl
from openpyxl.styles import Font

def generate_policy_comparison_data(output_file="policy_comparison_data.xlsx", count=100):
    """Generate XLSX file with policy comparison data"""
    print(f"Generating {count} policy comparison entries...")
    
    # Create a new workbook and select active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Comparison Data"
    
    # Add headers with bold formatting
    headers = ["Policy Number", "Source URL", "Destination URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Generate policy data
    policy_num = 573
    for i in range(count):
        row = i + 2  # Start at row 2 after headers
        
        # Create the URLs based on the policy number
        source_url = f"https://al-policies.exploremyplan.com/portal/web/medical-policies/-/mp-{policy_num}"
        dest_url = f"https://stage-us-mypolicies.itilitihealth.us/policy/938125692074/{policy_num}?lob=BCBS+AL"
        
        # Add data to the worksheet
        ws.cell(row=row, column=1, value=policy_num)
        ws.cell(row=row, column=2, value=source_url)
        ws.cell(row=row, column=3, value=dest_url)
    
    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 70
    
    # Save the workbook
    wb.save(output_file)
    print(f"Generated XLSX file: {output_file}")
    print(f"All entries use Policy Number: {policy_num}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate Policy Comparison XLSX Test Data')
    parser.add_argument('--output', default='policy_comparison_data.xlsx',
                        help='Output XLSX file name')
    parser.add_argument('--count', type=int, default=100,
                        help='Number of policy entries to generate')
    
    args = parser.parse_args()
    generate_policy_comparison_data(args.output, args.count)