import time
import re
import argparse
import os
import json
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import PatternFill, Font
import difflib

def get_text_from_source_url(url, headless=True):
    """
    Parser specifically for the source URL
    """
    print(f"Fetching source URL: {url}")
    
    # Set up Chrome options
    options = Options()
    if headless:
        options.add_argument('--headless')  # Run in background if headless=True
    
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--window-size=1920,1080')
    
    # Set a realistic user agent
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')
    
    # Initialize the driver - MODIFIED FOR GITHUB ACTIONS
    if os.environ.get('GITHUB_ACTIONS'):
        # Running in GitHub Actions
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        # Use system Chrome
        driver = webdriver.Chrome(options=options)
    else:
        # Running locally
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    try:
        # Load the page
        driver.get(url)
        
        # Wait for the page to load completely
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        
        # Give extra time for JavaScript content
        time.sleep(5)
        
        # Get the page source
        page_source = driver.page_source
        
        # Use BeautifulSoup to parse and clean
        soup = BeautifulSoup(page_source, 'html.parser')
        
        # UPDATED: Look for the div with class 'journal-content-article'
        # This matches the structure in the alabama.html file
        main_content = soup.find('div', class_='journal-content-article')
        
        if main_content:
            print(f"✅ Found main content using 'journal-content-article' class")
            # Remove specific unwanted elements
            for tag in main_content.find_all(["script", "style", "div", "a"], class_="hidden-print"):
                tag.decompose()
            
            # Get text from the cleaned main content
            text = main_content.get_text(separator='\n', strip=True)
            
            # Save a debug copy of the HTML content
            if not os.environ.get('GITHUB_ACTIONS'):  # Avoid writing debug files in GitHub Actions
                with open("source_html_debug.html", "w", encoding="utf-8") as f:
                    f.write(str(main_content))
                
        else:
            print(f"❌ Main content not found using 'journal-content-article'")
            # Try alternate class as fallback
            main_content = soup.find('div', class_='journey-content-article')
            
            if main_content:
                print(f"✅ Found main content using 'journey-content-article' class")
                # Remove specific unwanted elements
                for tag in main_content.find_all(["script", "style", "div", "a"], class_="hidden-print"):
                    tag.decompose()
                
                # Get text from the cleaned main content
                text = main_content.get_text(separator='\n', strip=True)
            else:
                print(f"❌ Main content not found for {url}")
                # If we can't find the specific content, fall back to the entire page
                # but still clean up unwanted elements
                for script_or_style in soup(['script', 'style', 'header', 'footer', 'nav']):
                    script_or_style.decompose()
                    
                text = soup.get_text(separator='\n', strip=True)
        
        # Clean up text
        text = re.sub(r'\n+', '\n', text)
        
        return text
    except Exception as e:
        print(f"Error with Selenium: {str(e)}")
        return f"Error fetching {url} with Selenium: {str(e)}"
    finally:
        # Always close the driver
        driver.quit()

def get_text_from_destination_url(url, headless=True):
    """
    Parser specifically for the destination URL in the main tag
    """
    print(f"Fetching destination URL: {url}")
    
    # Set up Chrome options
    options = Options()
    if headless:
        options.add_argument('--headless')  # Run in background if headless=True
    
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--window-size=1920,1080')
    
    # Set a realistic user agent
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
    options.add_argument(f'user-agent={user_agent}')
    
    # Initialize the driver - MODIFIED FOR GITHUB ACTIONS
    if os.environ.get('GITHUB_ACTIONS'):
        # Running in GitHub Actions
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        # Use system Chrome
        driver = webdriver.Chrome(options=options)
    else:
        # Running locally
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    try:
        # Load the page
        driver.get(url)
        
        # Wait for the page to load completely
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        
        # Give extra time for JavaScript content
        time.sleep(5)
        
        # Get the page source
        page_source = driver.page_source
        
        # Use BeautifulSoup to parse and clean
        soup = BeautifulSoup(page_source, 'html.parser')
        
        # SPECIFIC PARSING FOR DESTINATION URL - FIND MAIN TAG
        main_tag = soup.find('main')
        
        if main_tag:
            print(f"✅ Found main tag in destination URL")
            
            # Extract content from within the main tag
            content_wrapper_divs = main_tag.find_all('div', class_='contentWrapper')
            
            if content_wrapper_divs:
                all_text = []
                
                # Process each content wrapper div
                for div in content_wrapper_divs:
                    # Remove any scripts, styles, etc.
                    for element in div.find_all(['script', 'style']):
                        element.decompose()
                    
                    # Get the text from this div
                    div_text = div.get_text(separator='\n', strip=True)
                    all_text.append(div_text)
                
                # Combine all text from content wrappers
                text = '\n\n'.join(all_text)
            else:
                # If no content wrapper divs, get all text from main tag
                for element in main_tag.find_all(['script', 'style']):
                    element.decompose()
                text = main_tag.get_text(separator='\n', strip=True)
        else:
            print(f"❌ Main tag not found for {url}")
            # If we can't find the main tag, fall back to the entire page
            for script_or_style in soup(['script', 'style', 'header', 'footer', 'nav']):
                script_or_style.decompose()
                
            text = soup.get_text(separator='\n', strip=True)
        
        # Clean up text
        text = re.sub(r'\n+', '\n', text)
        
        return text
    except Exception as e:
        print(f"Error with Selenium: {str(e)}")
        return f"Error fetching {url} with Selenium: {str(e)}"
    finally:
        # Always close the driver
        driver.quit()

def extract_policy_number(url):
    """
    Extract policy number from URL
    """
    # For source URL format (e.g., https://al-policies.exploremyplan.com/portal/web/medical-policies/-/mp-573)
    match = re.search(r'/mp-(\d+)', url)
    if match:
        return match.group(1)
    
    # For destination URL format (e.g., https://stage-us-mypolicies.itilitihealth.us/policy/938125692074/573)
    match = re.search(r'/policy/\d+/(\d+)', url)
    if match:
        return match.group(1)
    
    # If no pattern matches, return 'Unknown'
    return 'Unknown'

def create_comparison_excel(url_pairs_results, output_file="policy_comparisons.xlsx"):
    """
    Create an Excel file with comparisons for multiple URL pairs
    """
    print(f"Creating Excel file with comparisons for {len(url_pairs_results)} policy pairs...")
    
    # Create the output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    
    # Create a new workbook or load existing one if it exists
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        # Find the last row to append new data
        last_row = ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Policy Comparisons"
        
        # Set up headers
        ws.cell(row=1, column=1, value="Policy Number").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Source URL").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Destination URL").font = Font(bold=True)
        ws.cell(row=1, column=4, value="Source Content").font = Font(bold=True)
        ws.cell(row=1, column=5, value="Destination Content").font = Font(bold=True)
        ws.cell(row=1, column=6, value="Status").font = Font(bold=True)
        ws.cell(row=1, column=7, value="Similarity Score").font = Font(bold=True)
        
        last_row = 1  # Start at row 2 (after headers)
    
    # Define fill colors
    match_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    source_only_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
    dest_only_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
    
    # Process each URL pair result
    for result in url_pairs_results:
        policy_number = result['policy_number']
        source_url = result['source_url']
        dest_url = result['dest_url']
        source_text = result['source_text']
        dest_text = result['dest_text']
        similarity = result['similarity']
        
        # Add policy info
        row = last_row + 1
        ws.cell(row=row, column=1, value=policy_number).font = Font(bold=True)
        ws.cell(row=row, column=2, value=source_url)
        ws.cell(row=row, column=3, value=dest_url)
        
        # Split text into lines
        source_lines = source_text.split('\n')
        dest_lines = dest_text.split('\n')
        
        # Use difflib to compare lines
        differ = difflib.Differ()
        diff = list(differ.compare(source_lines, dest_lines))
        
        # Process the differences
        for line in diff:
            prefix = line[0:2]
            content = line[2:]
            
            if prefix == '  ':  # Line in both files
                ws.cell(row=row, column=4, value=content)
                ws.cell(row=row, column=5, value=content)
                ws.cell(row=row, column=6, value="Match")
                
                # Apply white background
                ws.cell(row=row, column=4).fill = match_fill
                ws.cell(row=row, column=5).fill = match_fill
                
            elif prefix == '- ':  # Line only in source
                ws.cell(row=row, column=4, value=content)
                ws.cell(row=row, column=5, value="")
                ws.cell(row=row, column=6, value="Only in Source")
                
                # Apply light red background
                ws.cell(row=row, column=4).fill = source_only_fill
                
            elif prefix == '+ ':  # Line only in destination
                ws.cell(row=row, column=4, value="")
                ws.cell(row=row, column=5, value=content)
                ws.cell(row=row, column=6, value="Only in Destination")
                
                # Apply light green background
                ws.cell(row=row, column=5).fill = dest_only_fill
            
            # Add similarity score only for the first row of this policy
            if row == last_row + 1:
                ws.cell(row=row, column=7, value=f"{similarity:.2f}")
                
            row += 1
        
        # Update last_row for the next policy
        last_row = row
    
    # Add a legend at the end
    row = last_row + 2  # Add some space
    ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1, value="White")
    ws.cell(row=row, column=2, value="Content matches in both")
    ws.cell(row=row, column=1).fill = match_fill
    
    row += 1
    ws.cell(row=row, column=1, value="Light Red")
    ws.cell(row=row, column=2, value="Content only in Source")
    ws.cell(row=row, column=1).fill = source_only_fill
    
    row += 1
    ws.cell(row=row, column=1, value="Light Green")
    ws.cell(row=row, column=2, value="Content only in Destination")
    ws.cell(row=row, column=1).fill = dest_only_fill
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file updated: {output_file}")

# NEW FUNCTION: Load URL pairs from XLSX file
def load_url_pairs_from_xlsx(xlsx_path):
    """
    Load URL pairs from an XLSX file
    """
    url_pairs = []
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # Get the column indices (assuming first row contains headers)
        headers = [cell.value for cell in ws[1]]
        policy_col = headers.index("Policy Number") + 1 if "Policy Number" in headers else None
        source_col = headers.index("Source URL") + 1 if "Source URL" in headers else None
        dest_col = headers.index("Destination URL") + 1 if "Destination URL" in headers else None
        
        # Check if required columns exist
        if not all([source_col, dest_col]):
            print("Error: XLSX file must contain 'Source URL' and 'Destination URL' columns")
            return []
        
        # Read data from each row (skip header row)
        for row_idx in range(2, ws.max_row + 1):
            source_url = ws.cell(row=row_idx, column=source_col).value
            dest_url = ws.cell(row=row_idx, column=dest_col).value
            
            # Skip rows with missing URLs
            if not source_url or not dest_url:
                continue
            
            url_pairs.append({
                'source_url': source_url.strip() if isinstance(source_url, str) else source_url,
                'dest_url': dest_url.strip() if isinstance(dest_url, str) else dest_url
            })
        
        print(f"Loaded {len(url_pairs)} URL pairs from XLSX file")
        return url_pairs
        
    except Exception as e:
        print(f"Error loading XLSX file: {str(e)}")
        return []

# The original load_url_pairs_from_config function is kept for backward compatibility
def load_url_pairs_from_config(config_path):
    """
    Load URL pairs from a configuration file (JSON or text)
    """
    if config_path.endswith('.xlsx'):
        return load_url_pairs_from_xlsx(config_path)
    elif config_path.endswith('.json'):
        with open(config_path, 'r') as f:
            config = json.load(f)
            return config['url_pairs']
    else:
        # Assume it's a text file with source,destination format
        url_pairs = []
        with open(config_path, 'r') as f:
            for line in f:
                if line.strip() and ',' in line:
                    source, dest = line.strip().split(',', 1)
                    url_pairs.append({
                        'source_url': source.strip(),
                        'dest_url': dest.strip()
                    })
        return url_pairs

def compare_url_pairs(url_pairs, headless=True, max_pairs=None):
    """
    Compare multiple URL pairs and return the results
    Added max_pairs parameter to limit the number of pairs processed
    """
    results = []
    
    # Limit the number of pairs if specified
    if max_pairs is not None and max_pairs > 0:
        url_pairs = url_pairs[:max_pairs]
    
    for idx, pair in enumerate(url_pairs):
        source_url = pair['source_url']
        dest_url = pair['dest_url']
        policy_number = extract_policy_number(source_url)
        
        print(f"\nProcessing Policy #{policy_number} ({idx+1}/{len(url_pairs)})")
        print("=" * 80)
        
        # Fetch source content
        print("Fetching source URL...")
        source_text = get_text_from_source_url(source_url, headless)
        
        # Add a delay to avoid being rate-limited
        time.sleep(3)
        
        # Fetch destination content
        print("Fetching destination URL...")
        dest_text = get_text_from_destination_url(dest_url, headless)
        
        # Save the extracted text to files (optional)
        os.makedirs('output', exist_ok=True)
        with open(f"output/source_{policy_number}.txt", "w", encoding="utf-8") as f:
            f.write(source_text)
        
        with open(f"output/dest_{policy_number}.txt", "w", encoding="utf-8") as f:
            f.write(dest_text)
        
        # Calculate similarity if we have valid content
        if not source_text.startswith("Error") and not dest_text.startswith("Error"):
            # Simple word-based comparison
            source_words = set(source_text.lower().split())
            dest_words = set(dest_text.lower().split())
            
            common_words = source_words.intersection(dest_words)
            
            print(f"Source URL has approximately {len(source_words)} unique words")
            print(f"Destination URL has approximately {len(dest_words)} unique words")
            print(f"They share approximately {len(common_words)} unique words in common")
            
            # Calculate similarity
            similarity = len(common_words) / (len(source_words.union(dest_words)))
            print(f"Simple similarity score: {similarity:.2f}")
            
            results.append({
                'policy_number': policy_number,
                'source_url': source_url,
                'dest_url': dest_url,
                'source_text': source_text,
                'dest_text': dest_text,
                'similarity': similarity
            })
        else:
            if source_text.startswith("Error"):
                print(f"Couldn't process source URL: {source_text}")
            if dest_text.startswith("Error"):
                print(f"Couldn't process destination URL: {dest_text}")
            
            # Still add to results for tracking
            results.append({
                'policy_number': policy_number,
                'source_url': source_url,
                'dest_url': dest_url,
                'source_text': "Error fetching source URL" if source_text.startswith("Error") else source_text,
                'dest_text': "Error fetching destination URL" if dest_text.startswith("Error") else dest_text,
                'similarity': 0.0
            })
    
    return results

def main():
    parser = argparse.ArgumentParser(description='Compare text content from multiple website pairs')
    parser.add_argument('--headless', action='store_true', 
                        help='Run in headless mode (no browser UI)')
    parser.add_argument('--output', default='policy_comparisons.xlsx',
                        help='Output Excel file name')
    parser.add_argument('--config', default='url_pairs.xlsx',
                        help='Configuration file with URL pairs (XLSX, JSON, or CSV)')
    parser.add_argument('--max', type=int, default=0,
                        help='Maximum number of URL pairs to process (0 = all)')
    args = parser.parse_args()
    
    # Check if config file exists, if not create a sample XLSX
    if not os.path.exists(args.config):
        print(f"Configuration file {args.config} not found. Creating a sample file...")
        
        # Create a sample XLSX file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "URL Pairs"
        
        # Add headers
        ws.cell(row=1, column=1, value="Policy Number").font = Font(bold=True)
        ws.cell(row=1, column=2, value="Source URL").font = Font(bold=True)
        ws.cell(row=1, column=3, value="Destination URL").font = Font(bold=True)
        
        # Add sample data
        sample_data = [
            (573, "https://al-policies.exploremyplan.com/portal/web/medical-policies/-/mp-573", 
                  "https://stage-us-mypolicies.itilitihealth.us/policy/938125692074/573?lob=BCBS+AL"),
            (573, "https://al-policies.exploremyplan.com/portal/web/medical-policies/-/mp-573", 
                  "https://stage-us-mypolicies.itilitihealth.us/policy/938125692074/573?lob=BCBS+AL")
        ]
        
        for idx, (policy_num, source_url, dest_url) in enumerate(sample_data):
            row = idx + 2  # Start after header row
            ws.cell(row=row, column=1, value=policy_num)
            ws.cell(row=row, column=2, value=source_url)
            ws.cell(row=row, column=3, value=dest_url)
        
        # Create the directory if it doesn't exist
        os.makedirs(os.path.dirname(args.config) if os.path.dirname(args.config) else '.', exist_ok=True)
        
        # Save the workbook
        wb.save(args.config)
        
        print(f"Created sample configuration file {args.config}. Please edit it with your URL pairs and run again.")
        return
    
    # Load URL pairs from config
    url_pairs = load_url_pairs_from_config(args.config)
    print(f"Loaded {len(url_pairs)} URL pairs from {args.config}")
    
    # Compare URL pairs (with max limit if specified)
    max_pairs = args.max if args.max > 0 else None
    if max_pairs:
        print(f"Processing only the first {max_pairs} URL pairs")
    
    results = compare_url_pairs(url_pairs, args.headless, max_pairs)
    
    # Create or update Excel report
    create_comparison_excel(results, args.output)
    
    print(f"\nComparison complete. Results saved to {args.output}")

if __name__ == "__main__":
    main()